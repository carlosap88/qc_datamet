#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Stage 02: ingesta y saneamiento físico de archivos de entrada."""

from __future__ import annotations

import logging
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from time import perf_counter
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from qc_datamet.pipeline.base_stage import BaseStage
from qc_datamet.utils.logger import get_logger


class Stage02Ingestion(BaseStage):
    """
    Abre los archivos descubiertos por Stage01 y genera DataFrames.

    Stage02 realiza únicamente saneamiento físico:

    - limita la lectura a un rango seguro;
    - elimina filas completamente vacías;
    - elimina columnas completamente vacías;
    - ignora hojas completamente vacías;
    - conserva columnas con datos aunque no tengan encabezado.

    No valida todavía nombres, orden ni significado de las columnas.
    Esa responsabilidad corresponde a Stage03.
    """

    name = "stage02_ingestion"

    PROJECT_ROOT = Path(__file__).resolve().parents[3]

    REPORT_FILENAME = "02_Ingestion_Report.txt"

    ENGINE_BY_EXTENSION = {
        ".xlsx": "openpyxl",
        ".xlsm": "openpyxl",
        ".xls": "xlrd",
        ".xlsb": "pyxlsb",
    }

    # Margen adicional para encontrar columnas vacías intermedias o sobrantes.
    DEFAULT_READ_BUFFER_COLUMNS = 8

    # =========================================================================
    # VALIDACIÓN DE ENTRADA
    # =========================================================================

    def validate_inputs(self, data: Any) -> None:
        """Valida el inventario recibido desde Stage01."""

        if not isinstance(data, list):
            raise TypeError(
                "Stage02Ingestion debe recibir la lista de inventario "
                "generada por Stage01."
            )

        required_fields = {
            "source_type",
            "path",
            "relative_path",
            "filename",
            "extension",
        }

        for position, item in enumerate(data, start=1):
            if not isinstance(item, dict):
                raise TypeError(
                    f"El elemento {position} del inventario "
                    "no es un diccionario."
                )

            missing_fields = required_fields - set(item)

            if missing_fields:
                missing = ", ".join(sorted(missing_fields))

                raise ValueError(
                    f"El elemento {position} del inventario "
                    f"no contiene: {missing}"
                )

    # =========================================================================
    # EJECUCIÓN
    # =========================================================================

    def execute(
        self,
        data: Any = None,
    ) -> list[dict[str, Any]]:
        """Lee y sanea físicamente los archivos inventariados."""

        inventory: list[dict[str, Any]] = data
        settings = self.config.get("settings.yaml", {})

        logger = get_logger(
            name=self.name,
            settings=settings,
            project_root=self.PROJECT_ROOT,
        )

        expected_columns = self._get_expected_column_count()
        read_buffer = self._get_read_buffer_columns(settings)
        safe_read_columns = expected_columns + read_buffer

        logger.info(
            "Inicio de Stage02 Ingestion | archivos=%d | "
            "columnas_esperadas=%d | limite_lectura=%d",
            len(inventory),
            expected_columns,
            safe_read_columns,
        )

        results: list[dict[str, Any]] = []

        for item in inventory:
            source_type = str(
                item.get("source_type", "")
            ).strip().lower()

            if source_type != "excel":
                result = self._build_skipped_result(
                    inventory_item=item,
                    reason=(
                        f"Fuente no implementada todavía en Stage02: "
                        f"{source_type or 'desconocida'}"
                    ),
                )

                results.append(result)

                self.state.add_warning(result["error"])
                logger.warning(result["error"])
                continue

            result = self._read_excel_file(
                inventory_item=item,
                expected_columns=expected_columns,
                safe_read_columns=safe_read_columns,
                logger=logger,
            )

            results.append(result)

        statistics = self._register_statistics(results)

        report_path = self._save_report(
            results=results,
            statistics=statistics,
            settings=settings,
        )

        if report_path is not None:
            self.state.set_statistic(
                "stage02_report_path",
                str(report_path),
            )

        self._print_summary(
            results=results,
            statistics=statistics,
            report_path=report_path,
            settings=settings,
        )

        logger.info(
            "Fin de Stage02 Ingestion | procesados=%d | "
            "correctos=%d | fallidos=%d | omitidos=%d | "
            "hojas_vacias=%d",
            statistics["files_processed"],
            statistics["files_successful"],
            statistics["files_failed"],
            statistics["files_skipped"],
            statistics["empty_sheets_ignored"],
        )

        self._raise_if_required(
            results=results,
            settings=settings,
        )

        return results

    # =========================================================================
    # CONFIGURACIÓN
    # =========================================================================
    def _resolve_path(
        self,
        configured_path: str | Path,
    ) -> Path:
        """Resuelve una ruta respecto a la raíz del proyecto."""

        path = Path(configured_path)

        if not path.is_absolute():
            path = self.PROJECT_ROOT / path

        return path.resolve()
    
    def _get_expected_column_count(self) -> int:
        """Obtiene la cantidad de columnas útiles desde excel_schema.yaml."""

        schema_file = self.config.get("excel_schema.yaml", {})

        if not isinstance(schema_file, dict):
            raise TypeError(
                "excel_schema.yaml debe contener un diccionario."
            )

        schema = schema_file.get(
            "excel_schema",
            schema_file,
        )

        if not isinstance(schema, dict):
            raise TypeError(
                "La raíz 'excel_schema' debe ser un diccionario."
            )

        possible_values = (
            schema.get("total_columns_expected"),
            schema.get("expected_columns"),
            schema.get("column_count"),
        )

        column_count: int | None = None

        for value in possible_values:
            if value is None:
                continue

            try:
                column_count = int(value)
                break
            except (TypeError, ValueError):
                continue

        if column_count is None:
            columns = schema.get("columns")

            if isinstance(columns, (list, tuple)):
                column_count = len(columns)

            elif isinstance(columns, dict):
                column_count = len(columns)

        if column_count is None:
            column_count = 52

        if column_count <= 0:
            raise ValueError(
                "excel_schema.yaml: la cantidad esperada de columnas "
                "debe ser mayor que cero."
            )

        return column_count

    def _get_read_buffer_columns(
        self,
        settings: dict[str, Any],
    ) -> int:
        """Obtiene el margen de columnas adicionales para la lectura segura."""

        ingestion_config = (
            settings
            .get("ingestion", {})
            .get("excel", {})
        )

        value = ingestion_config.get(
            "read_buffer_columns",
            self.DEFAULT_READ_BUFFER_COLUMNS,
        )

        try:
            buffer_columns = int(value)
        except (TypeError, ValueError):
            buffer_columns = self.DEFAULT_READ_BUFFER_COLUMNS

        return max(buffer_columns, 1)

    # =========================================================================
    # LECTURA DE EXCEL
    # =========================================================================

    def _read_excel_file(
        self,
        inventory_item: dict[str, Any],
        expected_columns: int,
        safe_read_columns: int,
        logger: logging.Logger,
    ) -> dict[str, Any]:
        """Abre y procesa físicamente un archivo Excel."""

        started = perf_counter()

        file_path = Path(
            inventory_item["path"]
        ).resolve()

        extension = str(
            inventory_item["extension"]
        ).strip().lower()

        engine = self.ENGINE_BY_EXTENSION.get(extension)

        result = self._build_base_result(
            inventory_item=inventory_item,
            file_path=file_path,
            extension=extension,
            engine=engine,
            expected_columns=expected_columns,
            safe_read_columns=safe_read_columns,
        )

        if engine is None:
            message = (
                f"No existe un motor configurado para "
                f"la extensión '{extension}'."
            )

            return self._finish_with_error(
                result=result,
                message=message,
                started=started,
                logger=logger,
            )

        if not file_path.is_file():
            message = f"El archivo no existe: {file_path}"

            return self._finish_with_error(
                result=result,
                message=message,
                started=started,
                logger=logger,
            )

        logger.info(
            "Leyendo archivo | estación=%s | motor=%s | archivo=%s",
            result["station_id"],
            engine,
            file_path,
        )

        try:
            if extension in {".xlsx", ".xlsm"}:
                sheets, empty_sheets = self._read_openpyxl_workbook(
                    file_path=file_path,
                    safe_read_columns=safe_read_columns,
                    logger=logger,
                )

            else:
                sheets, empty_sheets = self._read_pandas_workbook(
                    file_path=file_path,
                    engine=engine,
                    safe_read_columns=safe_read_columns,
                    logger=logger,
                )

            result["sheets"] = sheets
            result["empty_sheets_ignored"] = empty_sheets
            result["sheet_count"] = len(sheets)

            result["total_rows_before_cleanup"] = sum(
                sheet["rows_before_cleanup"]
                for sheet in sheets
            )

            result["total_columns_before_cleanup"] = max(
                (
                    sheet["columns_before_cleanup"]
                    for sheet in sheets
                ),
                default=0,
            )

            result["empty_rows_removed"] = sum(
                sheet["empty_rows_removed"]
                for sheet in sheets
            )

            result["empty_columns_removed"] = sum(
                sheet["empty_columns_removed"]
                for sheet in sheets
            )

            result["total_rows"] = sum(
                sheet["rows"]
                for sheet in sheets
            )

            result["total_columns"] = max(
                (
                    sheet["columns"]
                    for sheet in sheets
                ),
                default=0,
            )

            result["headerless_columns_with_data"] = sum(
                len(sheet["headerless_columns_with_data"])
                for sheet in sheets
            )

            if not sheets:
                result["status"] = "skipped"
                result["error"] = (
                    "El archivo no contiene hojas con datos útiles."
                )

                self.state.add_warning(
                    f"{file_path.name}: {result['error']}"
                )

            else:
                result["status"] = "success"

        except MemoryError as error:
            message = (
                f"Memoria insuficiente leyendo {file_path}: {error}"
            )

            return self._finish_with_error(
                result=result,
                message=message,
                started=started,
                logger=logger,
            )

        except (
            ImportError,
            OSError,
            ValueError,
            PermissionError,
        ) as error:
            message = (
                f"No se pudo leer el archivo {file_path}: "
                f"{type(error).__name__}: {error}"
            )

            return self._finish_with_error(
                result=result,
                message=message,
                started=started,
                logger=logger,
            )

        except Exception as error:
            message = (
                f"Error inesperado leyendo {file_path}: "
                f"{type(error).__name__}: {error}"
            )

            return self._finish_with_error(
                result=result,
                message=message,
                started=started,
                logger=logger,
                include_traceback=True,
            )

        result["read_duration_seconds"] = round(
            perf_counter() - started,
            3,
        )

        return result

    def _read_openpyxl_workbook(
        self,
        file_path: Path,
        safe_read_columns: int,
        logger: logging.Logger,
    ) -> tuple[list[dict[str, Any]], int]:
        """Lee XLSX/XLSM en modo streaming y limita las columnas físicamente."""

        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
        )

        sheets: list[dict[str, Any]] = []
        empty_sheets_ignored = 0

        try:
            for worksheet in workbook.worksheets:
                sheet_started = perf_counter()

                declared_rows = int(
                    worksheet.max_row or 0
                )

                declared_columns = int(
                    worksheet.max_column or 0
                )

                if declared_columns > safe_read_columns:
                    logger.warning(
                        "Rango excesivo limitado | archivo=%s | "
                        "hoja=%s | columnas_declaradas=%d | "
                        "columnas_leidas=%d",
                        file_path.name,
                        worksheet.title,
                        declared_columns,
                        safe_read_columns,
                    )

                row_iterator = worksheet.iter_rows(
                    min_row=1,
                    max_row=declared_rows or None,
                    min_col=1,
                    max_col=safe_read_columns,
                    values_only=True,
                )

                dataframe = pd.DataFrame.from_records(
                    row_iterator
                )

                sheet_result = self._sanitize_dataframe(
                    dataframe=dataframe,
                    sheet_name=worksheet.title,
                    declared_rows=declared_rows,
                    declared_columns=declared_columns,
                    read_columns_limit=safe_read_columns,
                    read_duration_seconds=round(
                        perf_counter() - sheet_started,
                        3,
                    ),
                )

                if sheet_result is None:
                    empty_sheets_ignored += 1

                    logger.info(
                        "Hoja vacía ignorada | archivo=%s | hoja=%s",
                        file_path.name,
                        worksheet.title,
                    )
                    continue

                sheets.append(sheet_result)

                logger.info(
                    "Hoja leída | archivo=%s | hoja=%s | "
                    "filas=%d | columnas=%d | "
                    "filas_vacias_eliminadas=%d | "
                    "columnas_vacias_eliminadas=%d",
                    file_path.name,
                    worksheet.title,
                    sheet_result["rows"],
                    sheet_result["columns"],
                    sheet_result["empty_rows_removed"],
                    sheet_result["empty_columns_removed"],
                )

        finally:
            workbook.close()

        return sheets, empty_sheets_ignored

    def _read_pandas_workbook(
        self,
        file_path: Path,
        engine: str,
        safe_read_columns: int,
        logger: logging.Logger,
    ) -> tuple[list[dict[str, Any]], int]:
        """Lee formatos Excel distintos de XLSX mediante Pandas."""

        sheets: list[dict[str, Any]] = []
        empty_sheets_ignored = 0

        with pd.ExcelFile(
            file_path,
            engine=engine,
        ) as workbook:
            for sheet_name in workbook.sheet_names:
                sheet_started = perf_counter()

                dataframe = pd.read_excel(
                    workbook,
                    sheet_name=sheet_name,
                    header=None,
                    dtype=object,
                    usecols=range(safe_read_columns),
                )

                rows_before, columns_before = dataframe.shape

                sheet_result = self._sanitize_dataframe(
                    dataframe=dataframe,
                    sheet_name=sheet_name,
                    declared_rows=int(rows_before),
                    declared_columns=int(columns_before),
                    read_columns_limit=safe_read_columns,
                    read_duration_seconds=round(
                        perf_counter() - sheet_started,
                        3,
                    ),
                )

                if sheet_result is None:
                    empty_sheets_ignored += 1

                    logger.info(
                        "Hoja vacía ignorada | archivo=%s | hoja=%s",
                        file_path.name,
                        sheet_name,
                    )
                    continue

                sheets.append(sheet_result)

                logger.info(
                    "Hoja leída | archivo=%s | hoja=%s | "
                    "filas=%d | columnas=%d",
                    file_path.name,
                    sheet_name,
                    sheet_result["rows"],
                    sheet_result["columns"],
                )

        return sheets, empty_sheets_ignored

    # =========================================================================
    # SANEAMIENTO FÍSICO
    # =========================================================================

    def _sanitize_dataframe(
        self,
        dataframe: pd.DataFrame,
        sheet_name: str,
        declared_rows: int,
        declared_columns: int,
        read_columns_limit: int,
        read_duration_seconds: float,
    ) -> dict[str, Any] | None:
        """
        Elimina exclusivamente filas y columnas completamente vacías.

        Las columnas con datos y encabezado vacío se conservan. Stage03 será
        responsable de asignar un nombre temporal y validar su estructura.
        """

        rows_before, columns_before = dataframe.shape

        if dataframe.empty:
            return None

        dataframe = dataframe.replace(
            r"^\s*$",
            pd.NA,
            regex=True,
        )

        empty_row_mask = dataframe.isna().all(axis=1)
        empty_column_mask = dataframe.isna().all(axis=0)

        empty_rows_removed = int(
            empty_row_mask.sum()
        )

        empty_column_positions = [
            int(position) + 1
            for position, is_empty in enumerate(empty_column_mask)
            if bool(is_empty)
        ]

        empty_columns_removed = len(
            empty_column_positions
        )

        dataframe = dataframe.loc[
            ~empty_row_mask,
            ~empty_column_mask,
        ].copy()

        dataframe.reset_index(
            drop=True,
            inplace=True,
        )

        if dataframe.empty:
            return None

        rows_after, columns_after = dataframe.shape

        headerless_columns = self._find_headerless_columns_with_data(
            dataframe
        )

        return {
            "sheet_name": sheet_name,
            "declared_rows": declared_rows,
            "declared_columns": declared_columns,
            "read_columns_limit": read_columns_limit,
            "rows_before_cleanup": int(rows_before),
            "columns_before_cleanup": int(columns_before),
            "empty_rows_removed": empty_rows_removed,
            "empty_columns_removed": empty_columns_removed,
            "empty_column_positions": empty_column_positions,
            "rows": int(rows_after),
            "columns": int(columns_after),
            "headerless_columns_with_data": headerless_columns,
            "read_duration_seconds": read_duration_seconds,
            "dataframe": dataframe,
        }

    @staticmethod
    def _find_headerless_columns_with_data(
        dataframe: pd.DataFrame,
    ) -> list[dict[str, Any]]:
        """
        Detecta columnas cuyo encabezado está vacío pero contienen datos.

        No cambia el encabezado. Stage03 decidirá si debe asignarse un nombre
        temporal como UNKNOWN_01.
        """

        if dataframe.empty:
            return []

        header_row = dataframe.iloc[0]
        detected: list[dict[str, Any]] = []

        for position, column_label in enumerate(
            dataframe.columns,
            start=1,
        ):
            header_value = header_row[column_label]

            header_is_empty = (
                pd.isna(header_value)
                or str(header_value).strip() == ""
            )

            if not header_is_empty:
                continue

            values_below_header = dataframe.iloc[
                1:,
                dataframe.columns.get_loc(column_label),
            ]

            contains_data = bool(
                values_below_header.notna().any()
            )

            if contains_data:
                detected.append(
                    {
                        "position": position,
                        "excel_column": get_column_letter(position),
                        "temporary_name_suggestion": (
                            f"UNKNOWN_{position:02d}"
                        ),
                    }
                )

        return detected

    # =========================================================================
    # RESULTADOS
    # =========================================================================

    def _build_base_result(
        self,
        inventory_item: dict[str, Any],
        file_path: Path,
        extension: str,
        engine: str | None,
        expected_columns: int,
        safe_read_columns: int,
    ) -> dict[str, Any]:
        """Construye la estructura inicial del resultado de un archivo."""

        return {
            "source": inventory_item,
            "station_id": self._extract_station_id(
                inventory_item
            ),
            "path": str(file_path),
            "relative_path": inventory_item.get(
                "relative_path",
                file_path.name,
            ),
            "filename": inventory_item.get(
                "filename",
                file_path.name,
            ),
            "extension": extension,
            "engine": engine,
            "status": "pending",
            "expected_columns": expected_columns,
            "safe_read_columns": safe_read_columns,
            "sheet_count": 0,
            "empty_sheets_ignored": 0,
            "total_rows_before_cleanup": 0,
            "total_columns_before_cleanup": 0,
            "empty_rows_removed": 0,
            "empty_columns_removed": 0,
            "total_rows": 0,
            "total_columns": 0,
            "headerless_columns_with_data": 0,
            "read_duration_seconds": 0.0,
            "sheets": [],
            "error": None,
        }

    def _build_skipped_result(
        self,
        inventory_item: dict[str, Any],
        reason: str,
    ) -> dict[str, Any]:
        """Construye un resultado para una fuente no implementada."""

        file_path = Path(
            inventory_item.get("path", "")
        )

        return {
            "source": inventory_item,
            "station_id": self._extract_station_id(
                inventory_item
            ),
            "path": str(file_path),
            "relative_path": inventory_item.get(
                "relative_path",
                file_path.name,
            ),
            "filename": inventory_item.get(
                "filename",
                file_path.name,
            ),
            "extension": inventory_item.get(
                "extension",
                "",
            ),
            "engine": None,
            "status": "skipped",
            "expected_columns": 0,
            "safe_read_columns": 0,
            "sheet_count": 0,
            "empty_sheets_ignored": 0,
            "total_rows_before_cleanup": 0,
            "total_columns_before_cleanup": 0,
            "empty_rows_removed": 0,
            "empty_columns_removed": 0,
            "total_rows": 0,
            "total_columns": 0,
            "headerless_columns_with_data": 0,
            "read_duration_seconds": 0.0,
            "sheets": [],
            "error": reason,
        }

    def _finish_with_error(
        self,
        result: dict[str, Any],
        message: str,
        started: float,
        logger: logging.Logger,
        include_traceback: bool = False,
    ) -> dict[str, Any]:
        """Completa un resultado fallido y registra el error."""

        result["status"] = "error"
        result["error"] = message
        result["read_duration_seconds"] = round(
            perf_counter() - started,
            3,
        )

        self.state.add_error(message)

        if include_traceback:
            logger.exception(message)
        else:
            logger.error(message)

        return result

    def _extract_station_id(
        self,
        inventory_item: dict[str, Any],
    ) -> str:
        """Extrae la estación desde data/raw/<tipo>/<estación>/archivo."""

        relative_path = Path(
            inventory_item.get("relative_path", "")
        )

        parts = relative_path.parts
        source_type = str(
            inventory_item.get("source_type", "")
        ).lower()

        normalized_parts = [
            part.lower()
            for part in parts
        ]

        try:
            source_index = normalized_parts.index(
                source_type
            )

            station_id = parts[
                source_index + 1
            ]

            return str(station_id).strip().upper()

        except (ValueError, IndexError):
            return "UNKNOWN"

    # =========================================================================
    # ESTADÍSTICAS
    # =========================================================================

    def _register_statistics(
        self,
        results: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """Calcula y registra las estadísticas generales de Stage02."""

        status_counter = Counter(
            result["status"]
            for result in results
        )

        successful = [
            result
            for result in results
            if result["status"] == "success"
        ]

        statistics = {
            "files_processed": len(results),
            "files_successful": status_counter.get("success", 0),
            "files_failed": status_counter.get("error", 0),
            "files_skipped": status_counter.get("skipped", 0),
            "stations_found": len(
                {
                    result["station_id"]
                    for result in results
                    if result["station_id"] != "UNKNOWN"
                }
            ),
            "sheets_read": sum(
                result["sheet_count"]
                for result in successful
            ),
            "empty_sheets_ignored": sum(
                result["empty_sheets_ignored"]
                for result in results
            ),
            "rows_before_cleanup": sum(
                result["total_rows_before_cleanup"]
                for result in successful
            ),
            "empty_rows_removed": sum(
                result["empty_rows_removed"]
                for result in successful
            ),
            "rows_read": sum(
                result["total_rows"]
                for result in successful
            ),
            "empty_columns_removed": sum(
                result["empty_columns_removed"]
                for result in successful
            ),
            "headerless_columns_with_data": sum(
                result["headerless_columns_with_data"]
                for result in successful
            ),
            "read_duration_seconds": round(
                sum(
                    result["read_duration_seconds"]
                    for result in results
                ),
                3,
            ),
        }

        for name, value in statistics.items():
            self.state.set_statistic(
                f"stage02_{name}",
                value,
            )

        return statistics

    # =========================================================================
    # REPORTE
    # =========================================================================

    def _save_report(
        self,
        results: list[dict[str, Any]],
        statistics: dict[str, Any],
        settings: dict[str, Any],
    ) -> Path | None:
        """Genera el reporte técnico de ingesta."""

        reports_config = (
            self.config
            .get("reports.yaml", {})
            .get("reports", {})
        )

        general = reports_config.get(
            "general",
            {},
        )

        if not general.get("enabled", True):
            return None

        if not general.get(
            "generate_reports",
            True,
        ):
            return None

        configured_directory = (
            reports_config
            .get("directories", {})
            .get("import", "./reports/import")
        )

        report_directory = self._resolve_path(
            configured_directory
        )

        report_directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        report_path = (
            report_directory
            / self.REPORT_FILENAME
        )

        overwrite = (
            settings
            .get("execution", {})
            .get("overwrite_outputs", True)
        )

        if report_path.exists() and not overwrite:
            raise FileExistsError(
                f"El reporte ya existe: {report_path}"
            )

        project = settings.get(
            "project",
            {},
        )

        project_name = project.get(
            "name",
            "QC_DataMet",
        )

        project_version = project.get(
            "version",
            "0.1.0",
        )

        execution_date = datetime.now(
            timezone.utc
        ).strftime("%Y-%m-%d %H:%M:%S UTC")

        width = 118

        lines: list[str] = [
            "=" * width,
            f"{project_name} v{project_version}",
            "STAGE 02 - INGESTA Y SANEAMIENTO FÍSICO",
            "=" * width,
            "",
            f"Fecha de ejecución          : {execution_date}",
            "Estado                      : COMPLETADO",
            "",
            "RESUMEN GENERAL",
            "-" * width,
            (
                "Archivos procesados        : "
                f"{statistics['files_processed']}"
            ),
            (
                "Lecturas correctas         : "
                f"{statistics['files_successful']}"
            ),
            (
                "Lecturas fallidas          : "
                f"{statistics['files_failed']}"
            ),
            (
                "Archivos omitidos          : "
                f"{statistics['files_skipped']}"
            ),
            (
                "Estaciones identificadas   : "
                f"{statistics['stations_found']}"
            ),
            (
                "Hojas útiles               : "
                f"{statistics['sheets_read']}"
            ),
            (
                "Hojas vacías ignoradas     : "
                f"{statistics['empty_sheets_ignored']}"
            ),
            (
                "Filas antes de limpieza    : "
                f"{statistics['rows_before_cleanup']:,}"
            ),
            (
                "Filas vacías eliminadas    : "
                f"{statistics['empty_rows_removed']:,}"
            ),
            (
                "Filas resultantes          : "
                f"{statistics['rows_read']:,}"
            ),
            (
                "Columnas vacías eliminadas : "
                f"{statistics['empty_columns_removed']:,}"
            ),
            (
                "Columnas sin encabezado    : "
                f"{statistics['headerless_columns_with_data']:,}"
            ),
            (
                "Tiempo total de lectura    : "
                f"{statistics['read_duration_seconds']:.3f} s"
            ),
            "",
            "ARCHIVOS PROCESADOS",
            "-" * width,
            (
                f"{'N°':>3}  "
                f"{'ESTACIÓN':<10} "
                f"{'MOTOR':<10} "
                f"{'HOJAS':>6} "
                f"{'FILAS':>11} "
                f"{'COLUMNAS':>10} "
                f"{'F.VACÍAS':>10} "
                f"{'C.VACÍAS':>10} "
                f"{'ESTADO':<9} "
                "ARCHIVO"
            ),
            "-" * width,
        ]

        for index, result in enumerate(
            results,
            start=1,
        ):
            lines.append(
                f"{index:>3}  "
                f"{result['station_id']:<10} "
                f"{str(result['engine'] or '-'):10} "
                f"{result['sheet_count']:>6} "
                f"{result['total_rows']:>11,} "
                f"{result['total_columns']:>10} "
                f"{result['empty_rows_removed']:>10,} "
                f"{result['empty_columns_removed']:>10,} "
                f"{result['status'].upper():<9} "
                f"{result['relative_path']}"
            )

            for sheet in result["sheets"]:
                lines.extend(
                    [
                        (
                            f"     Hoja                 : "
                            f"{sheet['sheet_name']}"
                        ),
                        (
                            f"     Dimensión declarada  : "
                            f"{sheet['declared_rows']:,} × "
                            f"{sheet['declared_columns']:,}"
                        ),
                        (
                            f"     Dimensión resultante : "
                            f"{sheet['rows']:,} × "
                            f"{sheet['columns']:,}"
                        ),
                        (
                            f"     Columnas vacías      : "
                            f"{sheet['empty_column_positions']}"
                        ),
                        (
                            f"     Sin encabezado       : "
                            f"{sheet['headerless_columns_with_data']}"
                        ),
                    ]
                )

            if result["error"]:
                lines.append(
                    f"     Observación          : {result['error']}"
                )

            lines.append("-" * width)

        lines.extend(
            [
                "",
                "PRÓXIMA ETAPA",
                "-" * width,
                "Stage03 - Validación estructural",
                "",
                "Stage03 comparará los DataFrames resultantes con "
                "excel_schema.yaml:",
                "    - cantidad de columnas;",
                "    - nombres de encabezados;",
                "    - orden de columnas;",
                "    - columnas obligatorias;",
                "    - columnas faltantes o adicionales;",
                "    - columnas con datos y encabezado vacío.",
                "",
                "=" * width,
                "FIN DEL REPORTE",
                "=" * width,
            ]
        )

        report_path.write_text(
            "\n".join(lines),
            encoding="utf-8",
        )

        return report_path

    # =========================================================================
    # VALIDACIÓN DE SALIDA
    # =========================================================================

    def validate_outputs(
        self,
        result: Any,
    ) -> None:
        """Valida la salida generada por Stage02."""

        if not isinstance(result, list):
            raise TypeError(
                "Stage02Ingestion debe devolver una lista."
            )

        valid_statuses = {
            "success",
            "error",
            "skipped",
        }

        for position, item in enumerate(
            result,
            start=1,
        ):
            if not isinstance(item, dict):
                raise TypeError(
                    f"El resultado {position} no es un diccionario."
                )

            if item.get("status") not in valid_statuses:
                raise ValueError(
                    f"Estado inválido en el resultado "
                    f"{position}: {item.get('status')}"
                )

            for sheet in item.get("sheets", []):
                dataframe = sheet.get("dataframe")

                if not isinstance(
                    dataframe,
                    pd.DataFrame,
                ):
                    raise TypeError(
                        f"La hoja '{sheet.get('sheet_name')}' "
                        "no contiene un DataFrame válido."
                    )

    # =========================================================================
    # PRESENTACIÓN
    # =========================================================================

    def _print_summary(
        self,
        results: list[dict[str, Any]],
        statistics: dict[str, Any],
        report_path: Path | None,
        settings: dict[str, Any],
    ) -> None:
        """Muestra el resumen profesional de Stage02."""

        project = settings.get(
            "project",
            {},
        )

        project_name = project.get(
            "name",
            "QC_DataMet",
        )

        project_version = project.get(
            "version",
            "0.1.0",
        )

        width = 124

        print()
        print("=" * width)
        print(
            f"{project_name} v{project_version} — "
            "STAGE 02: INGESTA Y SANEAMIENTO FÍSICO"
        )
        print("=" * width)

        print(
            f"{'N°':>3}  "
            f"{'ESTACIÓN':<10} "
            f"{'MOTOR':<10} "
            f"{'HOJAS':>6} "
            f"{'FILAS':>11} "
            f"{'COLUMNAS':>10} "
            f"{'F.VACÍAS':>10} "
            f"{'C.VACÍAS':>10} "
            f"{'TIEMPO':>10} "
            f"{'ESTADO':<9} "
            "ARCHIVO"
        )

        print("-" * width)

        for index, result in enumerate(
            results,
            start=1,
        ):
            print(
                f"{index:>3}  "
                f"{result['station_id']:<10} "
                f"{str(result['engine'] or '-'):10} "
                f"{result['sheet_count']:>6} "
                f"{result['total_rows']:>11,} "
                f"{result['total_columns']:>10} "
                f"{result['empty_rows_removed']:>10,} "
                f"{result['empty_columns_removed']:>10,} "
                f"{result['read_duration_seconds']:>8.2f} s "
                f"{result['status'].upper():<9} "
                f"{result['relative_path']}"
            )

        print("-" * width)
        print(
            f"Archivos procesados        : "
            f"{statistics['files_processed']}"
        )
        print(
            f"Lecturas correctas         : "
            f"{statistics['files_successful']}"
        )
        print(
            f"Lecturas fallidas          : "
            f"{statistics['files_failed']}"
        )
        print(
            f"Archivos omitidos          : "
            f"{statistics['files_skipped']}"
        )
        print(
            f"Estaciones identificadas   : "
            f"{statistics['stations_found']}"
        )
        print(
            f"Hojas vacías ignoradas     : "
            f"{statistics['empty_sheets_ignored']}"
        )
        print(
            f"Filas antes de limpieza    : "
            f"{statistics['rows_before_cleanup']:,}"
        )
        print(
            f"Filas vacías eliminadas    : "
            f"{statistics['empty_rows_removed']:,}"
        )
        print(
            f"Filas resultantes          : "
            f"{statistics['rows_read']:,}"
        )
        print(
            f"Columnas vacías eliminadas : "
            f"{statistics['empty_columns_removed']:,}"
        )
        print(
            f"Columnas sin encabezado    : "
            f"{statistics['headerless_columns_with_data']:,}"
        )
        print(
            f"Tiempo total de lectura    : "
            f"{statistics['read_duration_seconds']:.3f} s"
        )
        print(
            f"Advertencias               : "
            f"{len(self.state.warnings)}"
        )
        print(
            f"Errores                    : "
            f"{len(self.state.errors)}"
        )
        print(
            f"Reporte técnico            : "
            f"{report_path or 'No generado'}"
        )
        print("=" * width)
        print()

    # =========================================================================
    # CONTROL DE ERRORES
    # =========================================================================

    def _raise_if_required(
        self,
        results: list[dict[str, Any]],
        settings: dict[str, Any],
    ) -> None:
        """Detiene el pipeline si existen errores y así está configurado."""

        failed_count = sum(
            result["status"] == "error"
            for result in results
        )

        stop_on_error = (
            settings
            .get("execution", {})
            .get("stop_on_error", True)
        )

        if failed_count and stop_on_error:
            raise RuntimeError(
                f"Stage02 terminó con {failed_count} "
                "archivo(s) que no pudieron leerse."
            )